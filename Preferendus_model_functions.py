# -*- coding: utf-8 -*-
"""
Created on Sun Jan  7 16:08:23 2024

@author: Sylvana

Functions file for the preferendus tool
"""
import sys
import os

# Get the parent directory
parent_dir = os.path.dirname(os.path.realpath(__file__))

# Add the parent directory to sys.path
sys.path.append(parent_dir)

import matplotlib.pyplot as plt
import numpy as np
from scipy.interpolate import pchip_interpolate
import xlwings as xw
from openpyxl.utils.cell import column_index_from_string, get_column_letter, coordinate_from_string
import pandas as pd
from decimal import Decimal

"""
Functions
"""  
def transform_list(input_list, numbers=[1.0, 2.0, 3.0, 4.0], letters=['D', 'C', 'B', 'A']):
    mapping = dict(zip(letters, numbers))
    reverse_mapping = dict(zip(numbers, letters))

    if isinstance(input_list[0], str):
        return [mapping[element] for element in input_list]
    elif isinstance(input_list[0], float):
        return [reverse_mapping[element] for element in input_list]
    else:
        return "Invalid input list"

def objective_values(working_tab, start, offset):
    """
    Function that takes a starting point and returns the values and preferences of a preference
    input:
        start = starting cell of the first variable
        offset = row offset of the variable of interest
    returns:
        objective_list = a list of the values and corresponding preferences of an objective (order: [val1, val2, val3, pref1, pref2, pref3])
    """ 
    objective_table = working_tab.range(start).offset(row_offset=offset).expand().value 
    objective_list = [item for sublist in zip(*objective_table) for item in sublist]
    
    if isinstance(objective_list[0], str):
        objective_list[0:4] = transform_list(objective_list[0:4])
    elif isinstance(objective_list[0], Decimal):
        objective_list[0:4] = (int(obj) for obj in objective_list[0:4])
    return objective_list

def var_values(working_tab, start):
    """
    Function that takes a starting point and returns the function values and of all variables in the variable table:
        start = starting cell of the first variable (x1)
    returns:
        var_list = a list of lists per preference with the values of the variable function
    """ 
    var_table = working_tab.range(start).expand().value 
    return var_table

def var_values_new(working_tab, start):
    """
    Function that takes a starting point and returns the function values and of all variables in the variable table:
        start = starting cell of the first variable (x1)
    returns:
        var_list = a list of lists per preference with the values of the variable function
    """ 
    var_table = working_tab.range(start).expand().value
    restructured_var_table = []

    # Iterate over the original list and group every four lists together
    for i in range(0, len(var_table), 4):
        group_of_lists = var_table[i:i+4]
        restructured_var_table.append(group_of_lists)
    return restructured_var_table

def weights_values(working_tab, start, offset):
    """
    Function that takes a starting point and returns the weights for every stakeholder:
        start = starting cell of the first variable
        offset = row offset of the variable of interest
    returns:
        weights_list = a list of the values and corresponding preferences of an objective (order: [w1, w2, w3, w4, w5])
    """ 
    weights_list = working_tab.range(start).offset(row_offset=offset).expand().value 
    return weights_list

def objective_formula_categorical(surface_list, preference_list, functions_list_new, x1, x2, x3, x4, x5, x_categorical):
    """
    Objective function to minimize/maximize a variable
    """
    total_houses = x1 + x2 + x3 + x4 + x5
    categorical_order = x_categorical - 1
    formula_list = []
    for j, i in enumerate(categorical_order):
        formula_categorical = (x1[j] * functions_list_new[i][0] + x2[j] * functions_list_new[i][1] + \
                               x3[j] * functions_list_new[i][2] + x4[j] * functions_list_new[i][3] + \
                               x5[j] * functions_list_new[i][4]) / total_houses[j]
        formula_list.append(formula_categorical)
    formula_array = np.array(formula_list)
    #print(formula_array)
    interpolation_vector = pchip_interpolate(preference_list[0:4],
                                             preference_list[4:8], 
                                             formula_array)
    #print(interpolation_vector)
    return formula_array, interpolation_vector

def objective_formula_categorical_differentiate(surface_list, preference_list, functions_list_new, x1, x2, x3, x4, x5, x1_categorical, x2_categorical, x3_categorical, x4_categorical, x5_categorical):
    """
    Objective function to minimize/maximize a variable
    """
    total_houses = x1 + x2 + x3 + x4 + x5
    x1_categorical_order = x1_categorical - 1
    x2_categorical_order = x2_categorical - 1
    x3_categorical_order = x3_categorical - 1
    x4_categorical_order = x4_categorical - 1
    x5_categorical_order = x5_categorical - 1
    formula_tuple_list = []
    formula_list = []
    for idx, (val1, val2, val3, val4, val5) in enumerate(zip(x1_categorical_order, x2_categorical_order, x3_categorical_order, x4_categorical_order, x5_categorical_order)):
        formula_tuple = (functions_list_new[val1][0], functions_list_new[val2][1] , functions_list_new[val3][2],  functions_list_new[val4][3], functions_list_new[val5][4])
        formula_categorical = (x1[idx] * functions_list_new[val1][0] + x2[idx] * functions_list_new[val2][1] + \
                               x3[idx] * functions_list_new[val3][2] + x4[idx] * functions_list_new[val4][3] + \
                               x5[idx] * functions_list_new[val5][4]) / total_houses[idx]
        formula_tuple_list.append(formula_tuple)
        formula_list.append(formula_categorical)
    formula_tuple_array = np.transpose(np.array(formula_tuple_list))
    formula_array = np.array(formula_list)
    #print(formula_array)
    interpolation_vector = pchip_interpolate(preference_list[0:4],
                                             preference_list[4:8], 
                                             formula_array)
    #print(interpolation_vector)
    return formula_tuple_array, interpolation_vector

def delta_balance_interpolation(preference_list, balance_list, interpolation_vector):
    """
    Intermediate function to calculate the delta balances for
    the categorical variables through interpolation
    """
    balance_vector_interpolation = pchip_interpolate(preference_list,
                                                     balance_list, 
                                                     interpolation_vector)
    return balance_vector_interpolation

def continuous_objective_formula_balance(i, objective_list, formula_list, balance_list, scenario_delta_balance_list, x1, x2, x3, x4, x5):
    """
    Objective function to minimize/maximize the balance variable
    """
    # First interpolate for the categorical variables of the stakeholder
    x1_obj1_balance_interpolation = delta_balance_interpolation(objective_list[i-4][0:4], scenario_delta_balance_list[0], formula_list[0][0])
    x1_obj2_balance_interpolation = delta_balance_interpolation(objective_list[i-3][0:4], scenario_delta_balance_list[1], formula_list[1][0])
    x1_obj3_balance_interpolation = delta_balance_interpolation(objective_list[i-2][0:4], scenario_delta_balance_list[2], formula_list[2][0])
    x1_obj4_balance_interpolation = delta_balance_interpolation(objective_list[i-1][0:4], scenario_delta_balance_list[3], formula_list[3][0])
    
    x2_obj1_balance_interpolation = delta_balance_interpolation(objective_list[i-4][0:4], scenario_delta_balance_list[4], formula_list[0][1])
    x2_obj2_balance_interpolation = delta_balance_interpolation(objective_list[i-3][0:4], scenario_delta_balance_list[5], formula_list[1][1])
    x2_obj3_balance_interpolation = delta_balance_interpolation(objective_list[i-2][0:4], scenario_delta_balance_list[6], formula_list[2][1])
    x2_obj4_balance_interpolation = delta_balance_interpolation(objective_list[i-1][0:4], scenario_delta_balance_list[7], formula_list[3][1])
    
    x3_obj1_balance_interpolation = delta_balance_interpolation(objective_list[i-4][0:4], scenario_delta_balance_list[8], formula_list[0][2])
    x3_obj2_balance_interpolation = delta_balance_interpolation(objective_list[i-3][0:4], scenario_delta_balance_list[9], formula_list[1][2])
    x3_obj3_balance_interpolation = delta_balance_interpolation(objective_list[i-2][0:4], scenario_delta_balance_list[10], formula_list[2][2])
    x3_obj4_balance_interpolation = delta_balance_interpolation(objective_list[i-1][0:4], scenario_delta_balance_list[11], formula_list[3][2])
    
    x4_obj1_balance_interpolation = delta_balance_interpolation(objective_list[i-4][0:4], scenario_delta_balance_list[12], formula_list[0][3])
    x4_obj2_balance_interpolation = delta_balance_interpolation(objective_list[i-3][0:4], scenario_delta_balance_list[13], formula_list[1][3])
    x4_obj3_balance_interpolation = delta_balance_interpolation(objective_list[i-2][0:4], scenario_delta_balance_list[14], formula_list[2][3])
    x4_obj4_balance_interpolation = delta_balance_interpolation(objective_list[i-1][0:4], scenario_delta_balance_list[15], formula_list[3][3])
    
    x5_obj1_balance_interpolation = delta_balance_interpolation(objective_list[i-4][0:4], scenario_delta_balance_list[16], formula_list[0][4])
    x5_obj2_balance_interpolation = delta_balance_interpolation(objective_list[i-3][0:4], scenario_delta_balance_list[17], formula_list[1][4])
    x5_obj3_balance_interpolation = delta_balance_interpolation(objective_list[i-2][0:4], scenario_delta_balance_list[18], formula_list[2][4])
    x5_obj4_balance_interpolation = delta_balance_interpolation(objective_list[i-1][0:4], scenario_delta_balance_list[19], formula_list[3][4])
    
    # Use these interpolations to calculate the saldo for each housing type
    saldo_x1 = x1 * (balance_list[0]+x1_obj1_balance_interpolation+x1_obj2_balance_interpolation+ \
                      x1_obj3_balance_interpolation+x1_obj4_balance_interpolation)
    saldo_x2 = x2 * (balance_list[1]+x2_obj1_balance_interpolation+x2_obj2_balance_interpolation+ \
                      x2_obj3_balance_interpolation+x2_obj4_balance_interpolation)
    saldo_x3 = x3 * (balance_list[2]+x3_obj1_balance_interpolation+x3_obj2_balance_interpolation+ \
                      x3_obj3_balance_interpolation+x3_obj4_balance_interpolation)
    saldo_x4 = x4 * (balance_list[3]+x4_obj1_balance_interpolation+x4_obj2_balance_interpolation+ \
                      x4_obj3_balance_interpolation+x4_obj4_balance_interpolation)
    saldo_x5 = x5 * (balance_list[4]+x5_obj1_balance_interpolation+x5_obj2_balance_interpolation+ \
                      x5_obj3_balance_interpolation+x5_obj4_balance_interpolation)
        
    # Retrieve the saldo formula by summing over the saldo for each housing type
    saldo_formula = saldo_x1+saldo_x2+saldo_x3+saldo_x4+saldo_x5
    # print(saldo_formula)
    interpolation_vector = pchip_interpolate(objective_list[i][0:4],
                                             objective_list[i][4:8], 
                                             saldo_formula)
    return interpolation_vector

def categorical_objective_formula_balance(i, objective_list, categorical_balance_sum_list, balance_list, x1, x2, x3, x4, x5):
    """
    Objective function to minimize/maximize the balance variable
    """
   
    # Calculate the saldo for each housing type using the null balance and the balance sum lists
    saldo_x1 = x1 * (balance_list[0] + categorical_balance_sum_list[0])
    saldo_x2 = x2 * (balance_list[1] + categorical_balance_sum_list[1])
    saldo_x3 = x3 * (balance_list[2] + categorical_balance_sum_list[2])
    saldo_x4 = x4 * (balance_list[3] + categorical_balance_sum_list[3])
    saldo_x5 = x5 * (balance_list[4] + categorical_balance_sum_list[4])
        
    # Retrieve the saldo formula by summing over the saldo for each housing type
    saldo_formula = saldo_x1+saldo_x2+saldo_x3+saldo_x4+saldo_x5
    # print(saldo_formula)
    interpolation_vector = pchip_interpolate(objective_list[i][0:4],
                                             objective_list[i][4:8], 
                                             saldo_formula)
    #print(objective_list[i])
    return interpolation_vector
    
def constraint_max(variables):
    """Constraint that checks if the sum of x1 till x5 is not higher than 81 houses.

    :param variables: ndarray of n-by-m, with n the population size of the GA and m the number of variables.
    :return: list with scores of the constraint
    """
    x1 = variables[:, 0] 
    x2 = variables[:, 1]
    x3 = variables[:, 2]
    x4 = variables[:, 3]
    x5 = variables[:, 4]
    #x6 = variables[:, 5] now greenery
    
    return (x1 + x2 + x3 + x4 + x5) - 81  # < 0

def constraint_min(variables):
    """Constraint that checks if the sum of x1 till x5 is not lower than 35 houses.

    :param variables: ndarray of n-by-m, with n the population size of the GA and m the number of variables.
    :return: list with scores of the constraint
    """
    x1 = variables[:, 0] 
    x2 = variables[:, 1]
    x3 = variables[:, 2]
    x4 = variables[:, 3]
    x5 = variables[:, 4]
    #x6 = variables[:, 5] now greenery

    return 35 - (x1 + x2 + x3 + x4 + x5)  # < 0

def constraint_affordable(variables):
    """Constraint that checks if the sum of x2 and x4 is not lower than 54 houses.

    :param variables: ndarray of n-by-m, with n the population size of the GA and m the number of variables.
    :return: list with scores of the constraint
    """
    x2 = variables[:, 1]
    x4 = variables[:, 3]

    return 54 - (x2 + x4)  # < 0

def create_linspaces(obj_list):
    """Create a list of arrays to use for plotting the preference curves

    :param variables: the objective list.
    :return: list with arrays
    """
    linspaces=[]
    for objective in obj_list:
        linspace = np.linspace(objective[0], objective[3])
        linspaces.append(linspace)
    return linspaces
    
def create_preference_arrays(obj_list, linspaces):
    """Create a list of arrays to use for plotting the preference curves

    :param variables: the objective list and arrays list.
    :return: interpolated list with arrays
    """    
    preference_functions = []
    i=0
    for objective in obj_list:
        preference_function = pchip_interpolate(objective[0:4], objective[4:8], linspaces[i])
        preference_functions.append(preference_function)
        i+=1
    return preference_functions

def create_formula_value_categorical(o, functions_list, design_variables_IMAP):
    """Create the formula value of categorical objectives for plotting the preference curves

    :param variables: the objective list.
    :return: list with arrays
    """
    total_houses = (design_variables_IMAP[0] + design_variables_IMAP[1] +
                    design_variables_IMAP[2] + design_variables_IMAP[3] + 
                    design_variables_IMAP[4])
    j = design_variables_IMAP[o+5] - 1
    formula_value = (functions_list[j][0] * design_variables_IMAP[0] + functions_list[j][1] * design_variables_IMAP[1] + \
                     functions_list[j][2] * design_variables_IMAP[2] + functions_list[j][3] * design_variables_IMAP[3] + \
                     functions_list[j][4] * design_variables_IMAP[4]) / total_houses
    return formula_value

def create_formula_value_saldo(i, objective_list, scenario_delta_balance_list, formula_list, functions_null_scenario_balance_list, design_variables_IMAP):
    """Create the formula value of the saldo for plotting the preference curves

    :param variables: the objective list.
    :return: list with arrays
    """
    # First interpolate for the categorical variables of the stakeholder
    x1_obj1_balance = delta_balance_interpolation(objective_list[i-4][0:4], scenario_delta_balance_list[0], formula_list[0])
    x1_obj2_balance = delta_balance_interpolation(objective_list[i-3][0:4], scenario_delta_balance_list[1], formula_list[1])
    x1_obj3_balance = delta_balance_interpolation(objective_list[i-2][0:4], scenario_delta_balance_list[2], formula_list[2])
    x1_obj4_balance = delta_balance_interpolation(objective_list[i-1][0:4], scenario_delta_balance_list[3], formula_list[3])
    
    x2_obj1_balance = delta_balance_interpolation(objective_list[i-4][0:4], scenario_delta_balance_list[4], formula_list[0])
    x2_obj2_balance = delta_balance_interpolation(objective_list[i-3][0:4], scenario_delta_balance_list[5], formula_list[1])
    x2_obj3_balance = delta_balance_interpolation(objective_list[i-2][0:4], scenario_delta_balance_list[6], formula_list[2])
    x2_obj4_balance = delta_balance_interpolation(objective_list[i-1][0:4], scenario_delta_balance_list[7], formula_list[3])
    
    x3_obj1_balance = delta_balance_interpolation(objective_list[i-4][0:4], scenario_delta_balance_list[8], formula_list[0])
    x3_obj2_balance = delta_balance_interpolation(objective_list[i-3][0:4], scenario_delta_balance_list[9], formula_list[1])
    x3_obj3_balance = delta_balance_interpolation(objective_list[i-2][0:4], scenario_delta_balance_list[10], formula_list[2])
    x3_obj4_balance = delta_balance_interpolation(objective_list[i-1][0:4], scenario_delta_balance_list[11], formula_list[3])
    
    x4_obj1_balance = delta_balance_interpolation(objective_list[i-4][0:4], scenario_delta_balance_list[12], formula_list[0])
    x4_obj2_balance = delta_balance_interpolation(objective_list[i-3][0:4], scenario_delta_balance_list[13], formula_list[1])
    x4_obj3_balance = delta_balance_interpolation(objective_list[i-2][0:4], scenario_delta_balance_list[14], formula_list[2])
    x4_obj4_balance = delta_balance_interpolation(objective_list[i-1][0:4], scenario_delta_balance_list[15], formula_list[3])
    
    x5_obj1_balance = delta_balance_interpolation(objective_list[i-4][0:4], scenario_delta_balance_list[16], formula_list[0])
    x5_obj2_balance = delta_balance_interpolation(objective_list[i-3][0:4], scenario_delta_balance_list[17], formula_list[1])
    x5_obj3_balance = delta_balance_interpolation(objective_list[i-2][0:4], scenario_delta_balance_list[18], formula_list[2])
    x5_obj4_balance = delta_balance_interpolation(objective_list[i-1][0:4], scenario_delta_balance_list[19], formula_list[3])
    
    # Use these interpolations to calculate the saldo for each housing type
    saldo_x1 = design_variables_IMAP[0] * (functions_null_scenario_balance_list[0]+x1_obj1_balance+x1_obj2_balance+x1_obj3_balance+x1_obj4_balance)
    saldo_x2 = design_variables_IMAP[1] * (functions_null_scenario_balance_list[1]+x2_obj1_balance+x2_obj2_balance+x2_obj3_balance+x2_obj4_balance)
    saldo_x3 = design_variables_IMAP[2] * (functions_null_scenario_balance_list[2]+x3_obj1_balance+x3_obj2_balance+x3_obj3_balance+x3_obj4_balance)
    saldo_x4 = design_variables_IMAP[3] * (functions_null_scenario_balance_list[3]+x4_obj1_balance+x4_obj2_balance+x4_obj3_balance+x4_obj4_balance)
    saldo_x5 = design_variables_IMAP[4] * (functions_null_scenario_balance_list[4]+x5_obj1_balance+x5_obj2_balance+x5_obj3_balance+x5_obj4_balance)
        
    # Retrieve the saldo formula by summing over the saldo for each housing type
    saldo_formula = saldo_x1+saldo_x2+saldo_x3+saldo_x4+saldo_x5
    # print(saldo_formula)
    formula_value = pchip_interpolate(objective_list[i][0:4],
                                             objective_list[i][4:8], 
                                             saldo_formula)
    return saldo_formula

def create_individual_preference_score(objective, formula_value):
    """Create the individual preference score for plotting the preference curves

    :param variables: the objective list.
    :return: list with arrays
    """
    individual_score = pchip_interpolate(objective[0:4], objective[4:8], formula_value)
    return individual_score
    

def create_save_categorical_plot(folder, linspace, preference_function, objective_outcome, individual_score, horizontal_offset, vertical_offset, o, s, i, working_tab, start_plot):
    fig = plt.figure()
    ax = fig.add_subplot(1, 1, 1)
    ax.plot(linspace, preference_function, label=f'Preference curve {i+1}')
    ax.scatter(objective_outcome, individual_score, label='Optimal solution IMAP', color='tab:purple')
    ax.set_xlim((0,5))
    plt.xticks([1, 2, 3, 4], ['D', 'C', 'B', 'A'])

    
    
    ax.set_ylim((0, 102))
    ax.set_title(f'objective {o+1} for stakeholder {s+1}')
    ax.set_xlabel('DCBA-score')
    ax.set_ylabel('Preference score')
    ax.grid()
    ax.legend()
    
    # image_name = f"Preference_curve_{i+1}.png"
    # image_path = os.path.join(folder, image_name)
    # plt.savefig(image_path)

    column_start, row_start = coordinate_from_string(start_plot)
    column_number = column_index_from_string(column_start) + o * horizontal_offset
    column_letter = get_column_letter(column_number)
    row_number = row_start + s * vertical_offset 
    location = f"{column_letter}{row_number}"
    print(location)
    working_tab.pictures.add(
            fig,
            name=f"Preference curve (s{s+1}_o{o+1})",
            update=True,
            left=working_tab.range(location).left,
            top=working_tab.range(location).top,
            height=200,
            width=300,
            )

def create_save_saldo_plot(folder, linspace, preference_function, objective_outcome, individual_score, horizontal_offset, vertical_offset, o, s, i, working_tab, start_plot):
    col_offset = o * horizontal_offset
    row_offset = s * vertical_offset
    fig = plt.figure()
    ax = fig.add_subplot(1, 1, 1)
    ax.plot(linspace, preference_function, label=f'Preference curve {i+1}')
    ax.scatter(objective_outcome, individual_score, label='Optimal solution IMAP', color='tab:purple')
    extension=1.1
    ax.set_xlim((0, linspace[-1]*extension))

    
    ax.set_ylim((0, 102))
    ax.set_title(f'objective {o+1} for stakeholder {s+1}')
    ax.set_xlabel('Saldo score')
    ax.set_ylabel('Preference score')
    ax.grid()
    ax.legend()
    
    # image_name = f"Preference_curve_{i+1}.png"
    # image_path = os.path.join(folder, image_name)
    # plt.savefig(image_path)

    column_start, row_start = coordinate_from_string(start_plot)
    column_number = column_index_from_string(column_start) + o * horizontal_offset
    column_letter = get_column_letter(column_number)
    row_number = row_start + s * vertical_offset 
    location = f"{column_letter}{row_number}"
    print(location)    
    working_tab.pictures.add(
        fig,
        name=f"Preference curve (s{s+1}_o{o+1})",
        update=True,
        left=working_tab.range(location).left,
        top=working_tab.range(location).top,
        height=200,
        width=300,
        )

# def constraint_space(variables):
#     """Constraint that checks if the sum of the build area is not higher than the maximum space.

#     :param variables: ndarray of n-by-m, with n the population size of the GA and m the number of variables.
#     :return: list with scores of the constraint
#     """
#     x1 = variables[:, 0] 
#     x2 = variables[:, 1]
#     x3 = variables[:, 2]
#     x4 = variables[:, 3]
#     x5 = variables[:, 4]
    
#     return (x1*surface_list[0])+(x2*surface_list[1])+(x3*surface_list[2])+(x4*surface_list[3])+(x5*surface_list[4]) - surface_list[5] < 0


# # Function that compares all alternatives
# @xw.func
# def tetra_function(weight,preferences):
#     score_tetra = TetraSolver().request(weight,preferences)
#     return score_tetra

# # Function that interpolates with the preferences, values and solutions of the objective formulas
# @xw.func
# def interpolate_excel(xi,yi,x):
#     try:
#         pchip_interpolate(xi,yi,x)
#     except:
#         print('Invalid input')
#     return pchip_interpolate(xi,yi,x)
 